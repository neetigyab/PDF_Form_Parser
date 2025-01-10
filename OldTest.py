import pdfplumber
import cv2
import numpy as np
from openpyxl import Workbook
import matplotlib.pyplot as plt
import cv2
import pytesseract

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Define the fields in the expected order
fields = [
    "Name",
    "Account number",
    "Last four digits of card",
    "Transaction date",
    "Amount",
    "Merchant name",
    "Transaction was not authorized",  # Checkbox entry
    "My card was",                     # Checkbox entry
    "Date you lost your card",
    "Time you lost your card",
    "Date you realised card was stolen",
    "Time you realised card was stolen",
    "Do you know who made the transaction",  # Checkbox entry
    "Have you given permission to anyone to use your card",  # Checkbox entry
    "When was the last time you used your card",
    "Last transaction amount",
    "Where do you normally store your card",
    "where do you normally store your PIN",
    "Other items that were stolen",
    "Have you filed police report",  # Checkbox entry
    "Officer name",
    "Report number",
    "Suspect name",
    "Date",
    "contact number",
    "Reason for dispute",
]

# Function to detect checkboxes dynamically using OpenCV
def detect_checkboxes_dynamic(image):
    # Convert to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Apply threshold to detect boxes
    _, thresh = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)

    # Find contours
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    checkbox_regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        aspect_ratio = w / float(h)

        # Heuristic for checkbox detection (approximate square size)
        if 0.8 <= aspect_ratio <= 1.5 and 10 <= w <= 100 and 10 <= h <= 100:
            checkbox_regions.append((x, y, w, h))

    return checkbox_regions

# Function to determine checkbox state and visualize
import cv2
import pytesseract

def process_and_visualize_checkboxes(image, checkbox_regions, page_number, lines_to_ignore):
    checkbox_states = {}
    output_image = image.copy()

    # Ignore the first few lines on the first page
    ignore_until_y = 0
    if page_number == 1:
        d = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT, config='--psm 6')
        line_positions = sorted([(d['top'][i], d['height'][i]) for i in range(len(d['level'])) if d['level'][i] == 5])
        if len(line_positions) >= lines_to_ignore:
            buffer = 10
            ignore_until_y = line_positions[lines_to_ignore - 1][0] + line_positions[lines_to_ignore - 1][1] + buffer

    checkbox_regions = sorted(checkbox_regions, key=lambda box: (box[1], box[0]))

    for i, (x, y, w, h) in enumerate(checkbox_regions):
        if page_number == 1 and y < ignore_until_y:
            continue

        checkbox_roi = image[y:y+h, x:x+w]
        gray = cv2.cvtColor(checkbox_roi, cv2.COLOR_BGR2GRAY)
        _, thresh = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)
        non_zero_count = cv2.countNonZero(thresh)
        total_pixels = w * h
        is_checked = non_zero_count > total_pixels * 0.5

        text_x_start = x + w + 4
        text_x_end = text_x_start + 120  # Adjust as needed
        text_roi = image[y:y+h, text_x_start:text_x_end]
        detected_text = pytesseract.image_to_string(text_roi, config='--psm 6').strip()
        detected_text = detected_text.split("YES",2)
        d = detected_text[0]

        print(f"Checkbox_{i} detected at ({x}, {y}, {w}, {h}) - Text: {d}")

        checkbox_states[f"Checkbox_{i}"] = "Checked" if is_checked else d or "No Text Found"

        color = (0, 255, 0) if is_checked else (0, 0, 255)
        cv2.rectangle(output_image, (x, y), (x+w, y+h), color, 2)

    return checkbox_states, output_image


# Function to extract and process text data from the PDF
def extract_text_and_checkboxes_dynamic(pdf_path, print_text=False):
    extracted_data = []
    checkbox_states = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            if not text:
                continue  # Skip pages without text

            # Extract the page as an image for OpenCV processing
            image = np.array(page.to_image().original)

            # Detect checkboxes dynamically
            checkbox_regions = detect_checkboxes_dynamic(image)

            # Determine checkbox states and visualize
            page_checkboxes, visualized_image = process_and_visualize_checkboxes(image, checkbox_regions, page_number, lines_to_ignore=10)
            checkbox_states.update(page_checkboxes)

            # Save the visualized image for debugging
            plt.imshow(cv2.cvtColor(visualized_image, cv2.COLOR_BGR2RGB))
            plt.title(f"Page {page_number} Checkboxes")
            plt.axis("off")
            plt.show()

            text_lines = text.split("\n")

            # Special handling for the 6th element
            if len(text_lines) > 6:
                text_lines = text_lines[:6] + text_lines[6].split() + text_lines[7:]

            # Split '11/27/2024 08:43:26 am + $329.98' into two parts
            if len(text_lines) > 12:
                parts = text_lines[12].split(" + ", 1)
                if len(parts) == 2:
                    text_lines[12:13] = parts

            # Remove specific unwanted elements
            for index in sorted([11, 10], reverse=True):
                if index < len(text_lines):
                    text_lines.pop(index)

            if page_number==2:
                if len(text_lines) > 2:
                    text_lines = text_lines[:2] + text_lines[2].split() + text_lines[3:]

                t = [' '.join(text_lines[6:8])]
                text_lines = text_lines[:2] + t

            if print_text:
                print(f"--- Page {page_number} Text ---")
                print(text_lines)
                print("-" * 50)

            extracted_data.extend(text_lines)

    return extracted_data, checkbox_states

# Function to assign text manually to fields, including checkboxes
def assign_text_to_fields_with_checkboxes(text_lines, checkboxes):
    """
    Assigns extracted text from text_lines and checkbox states to predefined fields based on page_number.

    Args:
        text_lines (list): The list of extracted text lines from the page.
        checkboxes (dict): A dictionary containing checkbox states and detected text.
        page_number (int): The current page number.

    Returns:
        assigned_data (dict): A dictionary mapping fields to extracted values.
    """
    # Predefined mapping based on PDF structure
    manual_mapping = {
        "Name": 0,
        "Account number": 1,
        "Last four digits of card": 2,
        "Transaction date": 3,
        "Amount": 4,
        "Merchant name": 5,
        "Date you lost your card": 6,
        "Time you lost your card": 7,
        "Date you realised card was stolen": 8,
        "Time you realised card was stolen": 9,
        "When was the last time you used your card": 10,
        "Last transaction amount": 11,
        "Where do you normally store your card": 12,
        "Where do you normally store your PIN": 13,
        "Other items that were stolen": None,  # Not found
        "Officer name": None,  # Not found
        "Report number": None,  # Not found
        "Suspect name": None,  # Not found
        "Date": 14,
        "Contact number": 15,
        "Reason for dispute": 16,
    }

    # Mapping for checkboxes on both pages
    checkbox_mappings = {
        21: "Transaction was not authorized",
        22: "My card was",
        23: "Do you know who made the transaction",
        24: "Have you given permission to anyone to use your card",
        0: "Have you filed police report",
    }

    # Initialize the assigned_data dictionary
    assigned_data = {}

    # Assign values based on manual mapping
    for field, index in manual_mapping.items():
        if isinstance(index, int) and index < len(text_lines):
            assigned_data[field] = text_lines[index].strip()
        elif isinstance(index, str):  # Checkbox data (for when it's predefined)
            assigned_data[field] = index.strip()
        else:
            assigned_data[field] = "n/a"  # Assign 'n/a' if no value exists

    # Now, assign checkbox text to fields based on the mappings
    for checkbox_index, label in checkbox_mappings.items():
        checkbox_text = checkboxes.get(f"Checkbox_{checkbox_index}", "No Text Found")
        assigned_data[label] = checkbox_text

    return assigned_data


# Function to write data to Excel using openpyxl
def write_to_excel(data, fields, output_path):
    wb = Workbook()
    sheet = wb.active

    # Write headers
    for col_index, field in enumerate(fields, start=1):
        sheet.cell(row=1, column=col_index, value=field)

    # Write data
    for col_index, field in enumerate(fields, start=1):
        sheet.cell(row=2, column=col_index, value=data.get(field, "n/a"))

    wb.save(output_path)

# Main processing
input_pdf = "ref\Dispute form new format check.pdf"
output_excel = "output_final.xlsx"

# Extract text data and checkboxes dynamically
text_data, checkboxes = extract_text_and_checkboxes_dynamic(input_pdf, print_text=True)

# Assign text data to fields
final_data = assign_text_to_fields_with_checkboxes(text_data, checkboxes)

# Write to Excel
write_to_excel(final_data, fields, output_excel)

print(f"Data has been written to {output_excel}")